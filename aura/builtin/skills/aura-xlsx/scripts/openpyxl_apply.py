#!/usr/bin/env python3
"""
Apply XLSX plan using openpyxl (supports structural edits).

This path may drop advanced features in some workbooks; the caller should inspect parts_diff.
"""
from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any, Dict, List

def _load_openpyxl():
    try:
        from openpyxl import load_workbook, Workbook  # type: ignore
        from openpyxl.utils.cell import range_boundaries  # type: ignore
        from openpyxl.formula.translate import Translator  # type: ignore
        from openpyxl.workbook.defined_name import DefinedName  # type: ignore
        return load_workbook, Workbook, range_boundaries, Translator, DefinedName
    except Exception as e:  # pragma: no cover
        raise SystemExit(f"openpyxl is required for this plan, but it is not available in this environment: {e}") from e

def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.lstrip().startswith("=")


def _guard_formula_overwrite(*, sheet_name: str, coord: str, existing: Any, new: Any, allow: bool) -> None:
    if allow:
        return
    if _is_formula(existing) and not _is_formula(new):
        raise ValueError(
            f"Refusing to overwrite formula cell {sheet_name}!{coord} with a non-formula value. "
            "Fill input/assumption cells instead, or write a formula string (starts with '='). "
            "If you truly need to replace formulas with values, set constraints.allow_overwrite_formula_cells=true."
        )


def _copy_cell_style(src, dst):
    dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.comment = src.comment
    dst.hyperlink = src.hyperlink

def _carry_row_style(ws, src_row: int, dst_row: int, min_col: int, max_col: int):
    for col in range(min_col, max_col + 1):
        s = ws.cell(row=src_row, column=col)
        d = ws.cell(row=dst_row, column=col)
        _copy_cell_style(s, d)
    # row height
    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

def _carry_col_style(ws, src_col: int, dst_col: int, min_row: int, max_row: int):
    for row in range(min_row, max_row + 1):
        s = ws.cell(row=row, column=src_col)
        d = ws.cell(row=row, column=dst_col)
        _copy_cell_style(s, d)
    try:
        ws.column_dimensions[ws.cell(row=1, column=dst_col).column_letter].width = ws.column_dimensions[ws.cell(row=1, column=src_col).column_letter].width
    except Exception:
        pass

def _apply_ops_openpyxl(*, wb, plan: Dict[str, Any], is_new: bool) -> List[Dict[str, Any]]:
    _load, _wb_cls, range_boundaries, Translator, DefinedName = _load_openpyxl()
    change_log: List[Dict[str, Any]] = []

    constraints = plan.get("constraints")
    allow_overwrite_formula_cells = bool(
        isinstance(constraints, dict) and constraints.get("allow_overwrite_formula_cells") is True
    )

    def ws(name: str):
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {name}")
        return wb[name]

    for op in plan.get("operations", []):
        kind = op.get("op")

        if kind == "create_workbook":
            if not is_new:
                raise ValueError("create_workbook is only allowed when creating a new workbook.")
            sheets = op.get("sheets") or []
            if sheets:
                wb.active.title = sheets[0]
                for name in sheets[1:]:
                    if name in wb.sheetnames:
                        continue
                    wb.create_sheet(title=name)
            active_name = op.get("active_sheet")
            if isinstance(active_name, str) and active_name in wb.sheetnames:
                wb.active = wb[active_name]
            change_log.append({"op": kind, "sheets": sheets})
            continue

        if kind == "add_sheet":
            name = op.get("name")
            if name in wb.sheetnames:
                raise ValueError(f"Sheet already exists: {name}")
            index = op.get("index")
            wb.create_sheet(title=name, index=int(index) if isinstance(index, int) else None)
            change_log.append({"op": kind, "name": name, "index": index})
            continue

        if kind == "rename_sheet":
            src = op.get("from")
            dst = op.get("to")
            if src not in wb.sheetnames:
                raise ValueError(f"Sheet not found: {src}")
            if dst in wb.sheetnames:
                raise ValueError(f"Target sheet already exists: {dst}")
            wb[src].title = dst
            change_log.append({"op": kind, "from": src, "to": dst})
            continue

        if kind == "remove_sheet":
            name = op.get("name")
            if name not in wb.sheetnames:
                raise ValueError(f"Sheet not found: {name}")
            if len(wb.sheetnames) <= 1:
                raise ValueError("Cannot remove the last remaining sheet.")
            wb.remove(wb[name])
            change_log.append({"op": kind, "name": name})
            continue

        if kind == "define_named_ranges":
            names = op.get("names", {})
            for dn_name, ref in names.items():
                if dn_name in wb.defined_names:
                    del wb.defined_names[dn_name]
                wb.defined_names.add(DefinedName(dn_name, attr_text=str(ref)))
            change_log.append({"op": kind, "count": len(names)})
            continue

        if kind == "freeze_panes":
            w = ws(op["sheet"])
            cell = op.get("cell", "A1")
            w.freeze_panes = cell
            change_log.append({"op": kind, "sheet": w.title, "cell": cell})
            continue

        if kind == "set_cells":
            w = ws(op["sheet"])
            for item in op.get("cells", []):
                coord = item["cell"]
                new_value = item.get("value")
                existing = w[coord].value
                _guard_formula_overwrite(
                    sheet_name=w.title,
                    coord=coord,
                    existing=existing,
                    new=new_value,
                    allow=allow_overwrite_formula_cells,
                )
                w[coord].value = new_value
            change_log.append({"op": kind, "sheet": w.title, "count": len(op.get("cells", []))})

        elif kind == "set_range":
            w = ws(op["sheet"])
            rng = op["range"]
            values = op.get("values", [])
            by_row = bool(op.get("by_row", True))
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            idx = 0
            if by_row:
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if idx >= len(values):
                            break
                        cell = w.cell(row=r, column=c)
                        new_value = values[idx]
                        _guard_formula_overwrite(
                            sheet_name=w.title,
                            coord=cell.coordinate,
                            existing=cell.value,
                            new=new_value,
                            allow=allow_overwrite_formula_cells,
                        )
                        cell.value = new_value
                        idx += 1
            else:
                for c in range(min_col, max_col + 1):
                    for r in range(min_row, max_row + 1):
                        if idx >= len(values):
                            break
                        cell = w.cell(row=r, column=c)
                        new_value = values[idx]
                        _guard_formula_overwrite(
                            sheet_name=w.title,
                            coord=cell.coordinate,
                            existing=cell.value,
                            new=new_value,
                            allow=allow_overwrite_formula_cells,
                        )
                        cell.value = new_value
                        idx += 1
            change_log.append({"op": kind, "sheet": w.title, "range": rng, "written": idx})

        elif kind == "fill_named_ranges":
            vals = op.get("values", {})
            wrote_total = 0
            for name, value in vals.items():
                dn = wb.defined_names.get(name)  # type: ignore
                if dn is None:
                    raise ValueError(f"Defined name not found: {name}")
                for sheet_name, ref in list(dn.destinations):
                    w = ws(sheet_name)
                    min_col, min_row, max_col, max_row = range_boundaries(ref)
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            cell = w.cell(row=r, column=c)
                            _guard_formula_overwrite(
                                sheet_name=w.title,
                                coord=cell.coordinate,
                                existing=cell.value,
                                new=value,
                                allow=allow_overwrite_formula_cells,
                            )
                            cell.value = value
                            wrote_total += 1
            change_log.append({"op": kind, "names": len(vals), "cells": wrote_total})

        elif kind == "insert_rows":
            w = ws(op["sheet"])
            idx = int(op["idx"])
            amount = int(op.get("amount", 1))
            carry_style = op.get("carry_style", "above")
            src_row = max(1, idx - 1)
            min_col = w.min_column or 1
            max_col = w.max_column or 1
            w.insert_rows(idx, amount)
            if carry_style == "above":
                for r in range(idx, idx + amount):
                    _carry_row_style(w, src_row, r, min_col, max_col)
            change_log.append({"op": kind, "sheet": w.title, "idx": idx, "amount": amount, "carry_style": carry_style})

        elif kind == "insert_cols":
            w = ws(op["sheet"])
            idx = int(op["idx"])
            amount = int(op.get("amount", 1))
            carry_style = op.get("carry_style", "left")
            src_col = max(1, idx - 1)
            min_row = w.min_row or 1
            max_row = w.max_row or 1
            w.insert_cols(idx, amount)
            if carry_style == "left":
                for c in range(idx, idx + amount):
                    _carry_col_style(w, src_col, c, min_row, max_row)
            change_log.append({"op": kind, "sheet": w.title, "idx": idx, "amount": amount, "carry_style": carry_style})

        elif kind == "copy_range":
            w = ws(op["sheet"])
            src = op["src"]
            dst = op["dst"]
            copy_style = bool(op.get("copy_style", True))
            min_col, min_row, max_col, max_row = range_boundaries(src)
            dst_col, dst_row, _, _ = range_boundaries(f"{dst}:{dst}")
            height = max_row - min_row + 1
            width = max_col - min_col + 1
            for r_off in range(height):
                for c_off in range(width):
                    s_cell = w.cell(row=min_row + r_off, column=min_col + c_off)
                    d_cell = w.cell(row=dst_row + r_off, column=dst_col + c_off)
                    v = s_cell.value
                    if isinstance(v, str) and v.startswith("="):
                        try:
                            d_cell.value = Translator(v, origin=s_cell.coordinate).translate_formula(d_cell.coordinate)
                        except Exception:
                            d_cell.value = v
                    else:
                        _guard_formula_overwrite(
                            sheet_name=w.title,
                            coord=d_cell.coordinate,
                            existing=d_cell.value,
                            new=v,
                            allow=allow_overwrite_formula_cells,
                        )
                        d_cell.value = v
                    if copy_style:
                        _copy_cell_style(s_cell, d_cell)
            change_log.append({"op": kind, "sheet": w.title, "src": src, "dst": dst, "copy_style": copy_style})

        elif kind == "append_table_rows":
            w = ws(op["sheet"])
            tbl = op["table"]
            header_row = int(tbl["header_row"])
            start_col = int(tbl["start_col"])
            end_col = int(tbl["end_col"])
            rows = op.get("rows", [])
            style_from = op.get("style_from", "last_row")  # last_row | header_row
            copy_formulas = bool(op.get("copy_formulas", True))

            # Build header -> column index mapping from header_row
            headers = {}
            for c in range(start_col, end_col + 1):
                txt = w.cell(row=header_row, column=c).value
                if txt is None:
                    continue
                headers[str(txt).strip()] = c

            # Find last used row within table columns (simple heuristic)
            last = header_row
            for r in range(header_row + 1, (w.max_row or header_row) + 1):
                any_val = False
                for c in range(start_col, end_col + 1):
                    if w.cell(row=r, column=c).value not in (None, ""):
                        any_val = True
                        break
                if any_val:
                    last = r

            template_row = last if style_from == "last_row" else header_row
            insert_at = last + 1
            amount = len(rows)
            if amount == 0:
                continue

            # Insert rows
            w.insert_rows(insert_at, amount)

            # Style + formula carry
            for i, row_obj in enumerate(rows):
                target_row = insert_at + i
                # copy style from template
                for c in range(start_col, end_col + 1):
                    _copy_cell_style(w.cell(row=template_row, column=c), w.cell(row=target_row, column=c))
                try:
                    w.row_dimensions[target_row].height = w.row_dimensions[template_row].height
                except Exception:
                    pass

                # copy formulas (translate relative references)
                if copy_formulas:
                    for c in range(start_col, end_col + 1):
                        v = w.cell(row=template_row, column=c).value
                        if isinstance(v, str) and v.startswith("="):
                            try:
                                origin = w.cell(row=template_row, column=c).coordinate
                                dest = w.cell(row=target_row, column=c).coordinate
                                w.cell(row=target_row, column=c).value = Translator(v, origin=origin).translate_formula(dest)
                            except Exception:
                                w.cell(row=target_row, column=c).value = v

                # fill values by header text if keys match, else by column letter
                for k, v in row_obj.items():
                    col_idx = headers.get(k)
                    if col_idx is None and isinstance(k, str) and len(k) <= 3 and k.isalpha():
                        # allow "A", "B" ...
                        col_idx = ord(k.upper()) - 64
                    if col_idx is not None:
                        w.cell(row=target_row, column=col_idx).value = v

            change_log.append({"op": kind, "sheet": w.title, "appended": amount, "at_row": insert_at})

        elif kind == "merge_cells":
            # Merge a range of cells
            w = ws(op["sheet"])
            rng = op["range"]
            w.merge_cells(rng)
            change_log.append({"op": kind, "sheet": w.title, "range": rng})

        elif kind == "set_style":
            # Set style (font, fill, alignment, border) for a cell or range
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            w = ws(op["sheet"])
            target = op.get("cell") or op.get("range")
            style_def = op.get("style", {})

            # Parse style definition
            font_def = style_def.get("font")
            fill_def = style_def.get("fill")
            align_def = style_def.get("alignment")
            border_def = style_def.get("border")
            number_format = style_def.get("number_format")

            # Build openpyxl style objects
            font_obj = None
            if font_def:
                font_obj = Font(
                    bold=font_def.get("bold", False),
                    italic=font_def.get("italic", False),
                    color=font_def.get("color"),
                    size=font_def.get("size"),
                    name=font_def.get("name"),
                )

            fill_obj = None
            if fill_def:
                fill_obj = PatternFill(
                    fill_type=fill_def.get("type", "solid"),
                    fgColor=fill_def.get("color"),
                )

            align_obj = None
            if align_def:
                align_obj = Alignment(
                    horizontal=align_def.get("horizontal"),
                    vertical=align_def.get("vertical"),
                    wrap_text=align_def.get("wrap_text", False),
                )

            border_obj = None
            if border_def:
                sides = {}
                for side_name in ["left", "right", "top", "bottom"]:
                    side_def = border_def.get(side_name)
                    if side_def:
                        sides[side_name] = Side(
                            style=side_def.get("style", "thin"),
                            color=side_def.get("color", "000000"),
                        )
                border_obj = Border(**sides)

            # Apply to cell or range
            if ":" in target:
                min_col, min_row, max_col, max_row = range_boundaries(target)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        cell = w.cell(row=r, column=c)
                        if font_obj:
                            cell.font = font_obj
                        if fill_obj:
                            cell.fill = fill_obj
                        if align_obj:
                            cell.alignment = align_obj
                        if border_obj:
                            cell.border = border_obj
                        if number_format:
                            cell.number_format = number_format
            else:
                cell = w[target]
                if font_obj:
                    cell.font = font_obj
                if fill_obj:
                    cell.fill = fill_obj
                if align_obj:
                    cell.alignment = align_obj
                if border_obj:
                    cell.border = border_obj
                if number_format:
                    cell.number_format = number_format

            change_log.append({"op": kind, "sheet": w.title, "target": target})

        elif kind == "set_row_height":
            w = ws(op["sheet"])
            row_idx = int(op["row"])
            height = float(op["height"])
            w.row_dimensions[row_idx].height = height
            change_log.append({"op": kind, "sheet": w.title, "row": row_idx, "height": height})

        elif kind == "set_column_width":
            w = ws(op["sheet"])
            col = op["column"]  # e.g., "A" or "B"
            width = float(op["width"])
            w.column_dimensions[col].width = width
            change_log.append({"op": kind, "sheet": w.title, "column": col, "width": width})

        else:
            raise ValueError(f"Unsupported op: {kind}")

    return change_log


def apply_plan_openpyxl(in_xlsx: Path, plan: Dict[str, Any], out_xlsx: Path) -> Dict[str, Any]:
    load_workbook, _wb_cls, _range_boundaries, _translator, _defined = _load_openpyxl()
    wb = load_workbook(filename=str(in_xlsx), data_only=False)
    change_log = _apply_ops_openpyxl(wb=wb, plan=plan, is_new=False)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_xlsx))
    wb.close()
    return {"changes": change_log}


def create_workbook_openpyxl(plan: Dict[str, Any], out_xlsx: Path) -> Dict[str, Any]:
    _load_workbook, Workbook, _range_boundaries, _translator, _defined = _load_openpyxl()
    wb = Workbook()
    change_log = _apply_ops_openpyxl(wb=wb, plan=plan, is_new=True)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_xlsx))
    wb.close()
    return {"changes": change_log}
